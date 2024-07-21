import re
import os
import gradio as gr
import codecs
from docx import Document
import pandas as pd
from openpyxl.styles import Alignment, Font, PatternFill
from tab4 import tab4_func as t4

def replace_special_periods(text):
    text = re.sub(r'\bDr\.', 'Dr<PERIOD>', text)
    text = re.sub(r'\.com\b', '<DOTCOM>', text)
    text = re.sub(r'\.org\b', '<DOTORG>', text)
    return text

def restore_special_periods(text):
    text = text.replace('Dr<PERIOD>', 'Dr.')
    text = text.replace('<DOTCOM>', '.com')
    text = text.replace('<DOTORG>', '.org')
    return text




def split_segment(segment, start_time, end_time):
    if start_time is None or end_time is None:
        return [(segment, start_time, end_time)]
    sentences = re.split(r'(?<!\.\.\.)\. ', segment)  # 三点リーダーはそのまま残す
    num_sentences = len(sentences)
    times = [(start_time + i * (end_time - start_time) / num_sentences) for i in range(num_sentences + 1)]
    return [(sentences[i], times[i], times[i+1]) for i in range(num_sentences)]

def merge_segments(segments):
    merged_segments = []
    buffer_segment = ""
    buffer_start = None
    buffer_end = None

    for segment, start, end in segments:
        if buffer_segment:
            buffer_segment += " " + segment
            buffer_end = end
            if segment.endswith('.'):
                merged_segments.append((buffer_segment, buffer_start, buffer_end))
                buffer_segment = ""
                buffer_start = None
                buffer_end = None
        else:
            if segment.endswith('.'):
                merged_segments.append((segment, start, end))
            else:
                buffer_segment = segment
                buffer_start = start
                buffer_end = end

    if buffer_segment:
        merged_segments.append((buffer_segment, buffer_start, buffer_end))

    return merged_segments

def process_vtt(lines):
    segments = []
    start_time = None
    end_time = None
    text = ""
    header = lines[0]

    # タイムスタンプの整形を適用
    lines = t4.unify_timestamps_forlist(lines, 'vtt')

    for line in lines[1:]:
        if re.match(r'^\d+$', line.strip()):
            if text:
                text = replace_special_periods(text)
                if start_time is not None and end_time is not None:
                    segments.extend(split_segment(text.strip(), start_time, end_time))
            text = ""
        elif '-->' in line:
            times = line.strip().split(' --> ')
            start_time = convert_time_to_seconds(times[0])
            end_time = convert_time_to_seconds(times[1])
        else:
            text += line.strip() + " "

    if text:
        text = replace_special_periods(text)
        if start_time is not None and end_time is not None:
            segments.extend(split_segment(text.strip(), start_time, end_time))

    merged_segments = merge_segments(segments)

    output = [header]
    segment_number = 1
    for text, start, end in merged_segments:
        text = restore_special_periods(text)
        output.append(f"{segment_number}")
        output.append(convert_seconds_to_time(start, 'vtt') + ' --> ' + convert_seconds_to_time(end, 'vtt'))
        output.append(text)
        segment_number += 1
        output.append("")  # セグメント間の改行を保持

    return '\n'.join(output)

def process_srt(lines):
    segments = []
    start_time = None
    end_time = None
    text = ""
    segment_index = 0

    # タイムスタンプの整形を適用
    lines = t4.unify_timestamps_forlist(lines, 'srt')

    for line in lines:
        if re.match(r'^\d+$', line.strip()):
            if text:
                text = replace_special_periods(text)
                segments.extend(split_segment(text.strip(), start_time, end_time))
            segment_index = int(line.strip())
            text = ""
        elif '-->' in line:
            times = line.strip().split(' --> ')
            start_time = convert_time_to_seconds(times[0].replace(',', '.'))
            end_time = convert_time_to_seconds(times[1].replace(',', '.'))
        else:
            text += line.strip() + " "

    if text:
        text = replace_special_periods(text)
        segments.extend(split_segment(text.strip(), start_time, end_time))

    merged_segments = merge_segments(segments)

    output = []
    segment_number = 0
    for text, start, end in merged_segments:
        text = restore_special_periods(text)
        output.append(f"{segment_number + 1}\n{convert_seconds_to_time(start,'srt').replace('.', ',')} --> {convert_seconds_to_time(end,'srt').replace('.', ',')}\n{text}\n")
        segment_number += 1

    return '\n'.join(output)



def convert_time_to_seconds(time_str):
    try:
        time_parts = time_str.split(':')
        if len(time_parts) == 3:
            h, m, s = time_parts
        elif len(time_parts) == 2:
            h = 0
            m, s = time_parts
        else:
            raise ValueError(f"Unexpected time format: {time_str}")
        
        h = float(h)
        m = float(m)
        s = float(s.replace(',', '.'))
        return h * 3600 + m * 60 + s
    except ValueError as e:
        raise ValueError(f"Error converting time: {time_str}, {e}")



def convert_seconds_to_time(seconds, format_type='vtt'):
    h = int(seconds // 3600)
    m = int((seconds % 3600) // 60)
    s = seconds % 60
    if format_type == 'vtt':
        return f"{h:01}:{m:02}:{s:06.3f}".replace(',', '.')
    else:
        return f"{h:02}:{m:02}:{s:06.3f}".replace('.', ',')
    
def process_file(input_file):
    if input_file is None:
        return None, None

    with open(input_file, 'r') as file:
        lines = file.readlines()

    _, file_extension = os.path.splitext(input_file)

    if file_extension.lower() == '.vtt':
        output = process_vtt(lines)
        output_file = os.path.splitext(input_file)[0] + '_edited.vtt'
    elif file_extension.lower() == '.srt':
        output = process_srt(lines)
        output_file = os.path.splitext(input_file)[0] + '_edited.srt'
    else:
        raise ValueError('Unsupported file format')

    with open(output_file, 'w') as file:
        file.write(output)

    # Add the output to a .docx file
    doc = Document()
    doc.add_paragraph(output)
    basename = os.path.splitext(input_file)[0]
    if file_extension.lower() == '.vtt':
        docx_file = basename + '_edited_vtt.docx'
    elif file_extension.lower() == '.srt':
        docx_file = basename + '_edited_srt.docx'
    else:
        pass
    doc.save(docx_file)

    output_html = f"""<head><meta charset="UTF-8"></head><body><pre style="white-space: pre-wrap; overflow-y: auto; height: 500px; word-wrap: break-word; padding: 10px; font-family: inherit; font-size: inherit;">{output}</pre></body>"""

    return output_html, [output_file, docx_file],output_file

##翻訳後の関数##
def correct_srt_format_from_text(text):
    # Remove all whitespace and newlines
    content = re.sub(r'[\u200B-\u200D\uFEFF]', '', text)
    content = re.sub(r'\s+', '', content)
    content = content.replace('\u200B', '')

    # Reconstruct the SRT format
    pattern = re.compile(r'(\d{1,4})(\d{2}:\d{2}:\d{2},\d{3}-->\d{2}:\d{2}:\d{2},\d{3})')
    #pattern = re.compile(r'(\d{1,4})\s*(\d{2}:\d{2}:\d{2},\d{3})\s*-->\s*(\d{2}:\d{2}:\d{2},\d{3})')
    matches = pattern.findall(content)
    
    segments = pattern.split(content)
    corrected_content = []
    
    for i in range(1, len(segments), 3):
        segment_id = segments[i]
        timestamp = segments[i + 1]
        text = segments[i + 2]
        
        corrected_content.append(f"{segment_id}")
        corrected_content.append(timestamp.replace('-->', ' --> '))
        corrected_content.append(text)

    # Ensure each subtitle block is separated by exactly one empty line
    final_content = "\n\n".join("\n".join(block) for block in zip(*[iter(corrected_content)]*3))
    
    return final_content


def correct_vtt_format_from_text(text):
    # Remove all whitespace and newlines
    content = re.sub(r'[\u200B-\u200D\uFEFF]', '', text)
    # Remove all whitespace and newlines
    content = re.sub(r'\s+', '', content)

    # Reconstruct the VTT format
    pattern = re.compile(r'(\d{1,4})(\d{1}:\d{2}:\d{2}\.\d{3}-->\d{1}:\d{2}:\d{2}\.\d{3})')
    #pattern = re.compile(r'(\d{1,4})(\d{1}:\d{2}:\d{2}\.\d{3})\s*-->\s*(\d{1}:\d{2}:\d{2}\.\d{3})')
    matches = pattern.findall(content)
    print(matches)
    if len(matches)==0:
        pattern = re.compile(r'(\d{1,4})(\d{2}:\d{2}:\d{2}\.\d{3}-->\d{1}:\d{2}:\d{2}\.\d{3})')
        matches = pattern.findall(content)
        print(matches)
        

    
    segments = pattern.split(content)
    corrected_content = []
    
    for i in range(1, len(segments), 3):
        segment_id = segments[i]
        timestamp = segments[i + 1]
        text = segments[i + 2]
        
        corrected_content.append(f"{segment_id}")
        corrected_content.append(timestamp.replace('-->', ' --> '))
        corrected_content.append(text)

    # Ensure each subtitle block is separated by exactly one empty line
    final_content = "WEBVTT\n\n" + "\n\n".join("\n".join(block) for block in zip(*[iter(corrected_content)]*3))
    
    return final_content

def vtt_translate(input_file, translated_content,output_file):
    if input_file==None or translated_content==None or output_file==None:
        return None
    
    ja_file_name, file_extension = os.path.splitext(input_file)
    output_ja_file_path = ja_file_name + "_edited_ja" + file_extension
    if file_extension==".srt":
        corrected_content=correct_srt_format_from_text(translated_content)
    
    elif file_extension==".vtt":
        corrected_content=correct_vtt_format_from_text(translated_content)
    
    with open(output_ja_file_path,'w',encoding='utf-8') as file:
            file.write(corrected_content+'\n')


    # excel出力
    output_excel_file = create_excel(output_file, output_ja_file_path)
    return [output_ja_file_path,output_excel_file]

##追加ぶん
def webvtt_remover(sentence): #tab3用。
    # 特殊文字を削除
    sentence = re.sub(r'[\u200B-\u200D\uFEFF]', '', sentence)
    
    # 改行文字を統一
    sentence = sentence.replace('\r\n', '\n').replace('\r', '\n')
    
    # VTTフォーマットを修正する正規表現パターン
    pattern = re.compile(r'(\d+)\n(\d{1}:\d{2}:\d{2}\.\d{3} --> \d{1}:\d{2}:\d{2}\.\d{3})', re.DOTALL)

    # 最初の一致部分を検索
    match = pattern.search(sentence)

    # 一致部分の開始インデックスを取得
    if match:
        start_index = match.start(1)  # 第1キャプチャグループの開始位置を取得
        rm_webvtt_sentence = sentence[start_index:]
    elif re.compile(r'(\d+)\n(\d{2}:\d{2}:\d{2}\.\d{3} --> \d{2}:\d{2}:\d{2}\.\d{3})', re.DOTALL).search(sentence):
        pattern = re.compile(r'(\d+)\n(\d{2}:\d{2}:\d{2}\.\d{3} --> \d{2}:\d{2}:\d{2}\.\d{3})', re.DOTALL)
        match = pattern.search(sentence)
        start_index = match.start(1)  # 第1キャプチャグループの開始位置を取得
        rm_webvtt_sentence = sentence[start_index:]
    else:
        rm_webvtt_sentence = sentence
    
    return rm_webvtt_sentence

def webvtt_rm(dic):

    sentence="".join(dic)
    # 改行文字を統一
    sentence = sentence.replace('\r\n', '\n').replace('\r', '\n')
    
    # VTTフォーマットを修正する正規表現パターン
    pattern = re.compile(r'(\d+)\n(\d{1}:\d{2}:\d{2}\.\d{3} --> \d{1}:\d{2}:\d{2}\.\d{3})', re.DOTALL)

    # 最初の一致部分を検索
    match = pattern.search(sentence)

    # 一致部分の開始インデックスを取得
    if match:
        start_index = match.start(1)  # 第1キャプチャグループの開始位置を取得
        rm_webvtt_sentence = sentence[start_index:]
    elif re.compile(r'(\d+)\n(\d{2}:\d{2}:\d{2}\.\d{3} --> \d{2}:\d{2}:\d{2}\.\d{3})', re.DOTALL).search(sentence):
        pattern = re.compile(r'(\d+)\n(\d{2}:\d{2}:\d{2}\.\d{3} --> \d{2}:\d{2}:\d{2}\.\d{3})', re.DOTALL)
        match = pattern.search(sentence)
        start_index = match.start(1)  # 第1キャプチャグループの開始位置を取得
        rm_webvtt_sentence = sentence[start_index:]
    else:
        rm_webvtt_sentence = sentence
    removed_dic=rm_webvtt_sentence.splitlines()

    return removed_dic

def create_excel(output_file, output_ja_file_path):
    segments = []
    if output_file.lower().endswith('.vtt') or output_file.lower().endswith('.srt'):
        with open(output_file, 'r') as f:
            lines = f.readlines()
            if output_file.lower().endswith('.vtt'):
                lines = webvtt_rm(lines)
            segments.extend(parse_segments(lines))
    
    if output_ja_file_path.lower().endswith('.vtt') or output_ja_file_path.lower().endswith('.srt'):
        with codecs.open(output_ja_file_path, 'r', 'utf-8') as f:
            ja_lines = f.readlines()
            if output_ja_file_path.lower().endswith('.vtt'):
                ja_lines = webvtt_rm(ja_lines)
            ja_segments = parse_segments(ja_lines)
    
    excel_data = []
    for (eng_segment, start, end), ja_segment in zip(segments, ja_segments):
        if start is not None and end is not None:
            excel_data.append({
                'ID': segments.index((eng_segment, start, end)) + 1,
                'Start': convert_seconds_to_time(start),
                'End': convert_seconds_to_time(end),
                'English Subtitle': eng_segment,
                'Japanese Subtitle': ja_segment[0]
            })
        else:
            print(f"Skipping segment due to missing time: {eng_segment}")

    df = pd.DataFrame(excel_data)
    output_excel_file = os.path.splitext(output_file)[0] + '.xlsx'
    df.to_excel(output_excel_file, index=False)

    with pd.ExcelWriter(output_excel_file, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='Subtitles')
        workbook = writer.book
        worksheet = writer.sheets['Subtitles']

        column_widths = {'A': 7, 'B': 25, 'C': 25, 'D': 90, 'E': 90}
        for column, width in column_widths.items():
            worksheet.column_dimensions[column].width = width

        for row in worksheet.iter_rows(min_row=2, max_row=len(df) + 1):
            for cell in row:
                if cell.column_letter == 'A':
                    cell.alignment = Alignment(horizontal='right', vertical='center')
                elif cell.column_letter in ['B', 'C']:
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                elif cell.column_letter in ['D', 'E']:
                    cell.alignment = Alignment(horizontal='left', vertical='center')

        for row in worksheet.iter_rows(min_row=2, max_row=len(df) + 1):
            worksheet.row_dimensions[row[0].row].height = 30

        header_font = Font(bold=True)
        for cell in worksheet["1:1"]:
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')
            cell.fill = PatternFill(start_color="DAEEF3", end_color="DAEEF3", fill_type="solid")

    return output_excel_file

def parse_segments(lines):
    segments = []
    start_time = None
    end_time = None
    text = ""
    for line in lines:
        if re.match(r'^\d+$', line.strip()):
            if text:
                segments.append((text.strip(), start_time, end_time))
            text = ""
        elif '-->' in line:
            times = line.strip().split(' --> ')
            try:
                start_time = convert_time_to_seconds(times[0])
                end_time = convert_time_to_seconds(times[1])
            except ValueError as e:
                print(f"Error parsing time: {times}, {e}")
                start_time, end_time = None, None
        else:
            text += line.strip() + " "
    if text:
        segments.append((text.strip(), start_time, end_time))
    return segments
